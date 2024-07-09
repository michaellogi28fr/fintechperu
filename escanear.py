import requests
from bs4 import BeautifulSoup
import openpyxl

# Creamos un nuevo libro de trabajo de Excel
workbook = openpyxl.Workbook()

# Seleccionamos la hoja activa
hoja_activa = workbook.active


url = requests.get('https://techcrunch.com/category/fintech/')
soup = BeautifulSoup(url.content, 'html.parser')


##-----------------------------------------------------------------------------------------------------


nombre = []

elementos_a = soup.find_all('h2', class_='has-link-color wp-block-post-title has-h-5-font-size wp-elements-565fa7bab0152bfdca0217543865c205')

for elemento_h2 in elementos_a:
    elemento_a = elemento_h2.find('a')
    if elemento_a:
        nombre.append(elemento_a.text.strip())


print(nombre)

##-----------------------------------------------------------------------------------------------------
precio = []

elementos_b = soup.find_all('div', class_='product-price-and-shipping')

for elemento_div in elementos_b:
    elemento_span = elemento_div.find('span')
    if elemento_span:
        precio.append(elemento_span.text.strip())

##-----------------------------------------------------------------------------------------------------


Todo_junto = [(nombre[i], precio[i]) for i in range(len(nombre))]
for producto in Todo_junto:
    print(producto[0], producto[1], '\n')


##-----------------------------------------------------------------------------------------------------


for i in range(len(nombre)):
    hoja_activa.cell(row=i+1, column=1, value=nombre[i])
    hoja_activa.cell(row=i+1, column=2, value=precio[i])

# Guardamos el libro de trabajo en un archivo
workbook.save("precios.xlsx")